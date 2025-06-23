"""
seofrog/exporters/excel_writer.py
Writer base para formatação Excel - compatível com pandas + openpyxl
"""

import pandas as pd
from typing import Optional

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from seofrog.utils.logger import get_logger

class ExcelWriter:
    """
    Writer base para formatação Excel
    Funciona com pandas + openpyxl para formatação avançada
    """
    
    def __init__(self):
        self.logger = get_logger('ExcelWriter')
        
        if not OPENPYXL_AVAILABLE:
            self.logger.warning("openpyxl não disponível - formatação limitada")
    
    def format_workbook(self, writer):
        """
        Aplica formatação profissional em todo o workbook
        """
        if not OPENPYXL_AVAILABLE:
            self.logger.debug("Formatação avançada não disponível - install openpyxl")
            return
        
        try:
            workbook = writer.book
            
            # Aplica formatação em todas as sheets
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                self._format_worksheet(worksheet)
                
            self.logger.debug(f"✅ Formatação aplicada em {len(workbook.sheetnames)} sheets")
                
        except Exception as e:
            self.logger.warning(f"Erro na formatação do workbook: {e}")
    
    def _format_worksheet(self, worksheet):
        """
        Formata uma worksheet individual com estilo profissional
        """
        try:
            # Skip sheets vazias
            if worksheet.max_row == 0:
                return
            
            # === FORMATAÇÃO DO HEADER (primeira linha) ===
            if worksheet.max_row > 0:
                for cell in worksheet[1]:
                    if cell.value:  # Só formata células com conteúdo
                        cell.font = Font(bold=True, color="FFFFFF", size=11)
                        cell.fill = PatternFill(
                            start_color="366092", 
                            end_color="366092", 
                            fill_type="solid"
                        )
                        cell.alignment = Alignment(
                            horizontal="center", 
                            vertical="center",
                            wrap_text=True
                        )
            
            # === AUTO-AJUSTE DE COLUNAS ===
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                # Calcula largura máxima da coluna
                for cell in column:
                    try:
                        cell_length = len(str(cell.value)) if cell.value else 0
                        if cell_length > max_length:
                            max_length = cell_length
                    except:
                        pass
                
                # Define largura (mínimo 8, máximo 50 para não quebrar layout)
                adjusted_width = min(max(max_length + 2, 8), 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # === CONGELA PRIMEIRA LINHA ===
            worksheet.freeze_panes = "A2"
            
            # === FORMATAÇÃO DE DADOS (linhas 2+) ===
            if worksheet.max_row > 1:
                for row in worksheet.iter_rows(min_row=2):
                    for cell in row:
                        if cell.value:
                            cell.alignment = Alignment(
                                horizontal="left", 
                                vertical="top",
                                wrap_text=False
                            )
            
        except Exception as e:
            self.logger.warning(f"Erro formatando worksheet: {e}")
    
    def apply_conditional_formatting(self, worksheet, column_letter: str, rule_type: str = "status"):
        """
        Aplica formatação condicional a uma coluna específica
        
        Args:
            worksheet: Worksheet do openpyxl
            column_letter: Letra da coluna (ex: 'C')
            rule_type: Tipo de regra ('status', 'performance', 'seo')
        """
        if not OPENPYXL_AVAILABLE:
            return
        
        try:
            from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
            from openpyxl.styles import PatternFill
            
            if rule_type == "status":
                # Formatação para status codes
                # Verde para 200, Vermelho para 4xx/5xx
                red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
                
                worksheet.conditional_formatting.add(
                    f"{column_letter}2:{column_letter}1000",
                    CellIsRule(operator="equal", formula=["200"], fill=green_fill)
                )
                worksheet.conditional_formatting.add(
                    f"{column_letter}2:{column_letter}1000", 
                    CellIsRule(operator="greaterThan", formula=["399"], fill=red_fill)
                )
                
            elif rule_type == "performance":
                # Escala de cores para response_time
                color_scale = ColorScaleRule(
                    start_type="num", start_value=0, start_color="00FF00",  # Verde
                    mid_type="num", mid_value=3, mid_color="FFFF00",       # Amarelo  
                    end_type="num", end_value=10, end_color="FF0000"       # Vermelho
                )
                worksheet.conditional_formatting.add(f"{column_letter}2:{column_letter}1000", color_scale)
                
        except Exception as e:
            self.logger.debug(f"Erro aplicando formatação condicional: {e}")
    
    def set_sheet_protection(self, worksheet, password: Optional[str] = None):
        """
        Protege a worksheet contra edição
        """
        if not OPENPYXL_AVAILABLE:
            return
        
        try:
            worksheet.protection.sheet = True
            if password:
                worksheet.protection.password = password
                
        except Exception as e:
            self.logger.debug(f"Erro protegendo worksheet: {e}")
    
    def add_data_validation(self, worksheet, cell_range: str, validation_type: str, formula: str):
        """
        Adiciona validação de dados a um range de células
        """
        if not OPENPYXL_AVAILABLE:
            return
        
        try:
            from openpyxl.worksheet.datavalidation import DataValidation
            
            dv = DataValidation(type=validation_type, formula1=formula)
            dv.add(cell_range)
            worksheet.add_data_validation(dv)
            
        except Exception as e:
            self.logger.debug(f"Erro adicionando validação: {e}")


# ==========================================
# FUNÇÕES AUXILIARES
# ==========================================

def create_formatted_writer(filepath: str, engine: str = "openpyxl") -> pd.ExcelWriter:
    """
    Cria um ExcelWriter pré-configurado
    
    Args:
        filepath: Caminho do arquivo
        engine: Engine do pandas ('openpyxl' ou 'xlsxwriter')
    
    Returns:
        pd.ExcelWriter configurado
    """
    if engine == "openpyxl" and not OPENPYXL_AVAILABLE:
        engine = "xlsxwriter"
        
    return pd.ExcelWriter(filepath, engine=engine)


def check_excel_dependencies() -> dict:
    """
    Verifica dependências disponíveis para Excel
    
    Returns:
        Dict com status das dependências
    """
    status = {
        'pandas': False,
        'openpyxl': False,
        'xlsxwriter': False
    }
    
    try:
        import pandas
        status['pandas'] = True
    except ImportError:
        pass
    
    try:
        import openpyxl
        status['openpyxl'] = True
    except ImportError:
        pass
    
    try:
        import xlsxwriter
        status['xlsxwriter'] = True
    except ImportError:
        pass
    
    return status